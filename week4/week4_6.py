import requests
from bs4 import BeautifulSoup
import openpyxl
keyword = input("검색어를 입력해주세요: ")
#wb# openpyexcel.Workbook() -> 새로운 workbook 열기
try:                                                    # 기존에 "navernews.xlsx" 얘가 있으면 여는 걸 시도해라
    wb = openpyxl.load_workbook("navernews.xlsx")       #sheet는 현재 활성화된 wb이다
    sheet = wb.active                                   #이 경우는 이미 헤더가 있음으로 굳이 해더를 쓰지 않아도 됨
    print("불러오기 완료")
#여기에서 navernews.xlsx -> 얜 상대 경로, 같은 폴더에 있는 파일만 ㄱㄴ
#절대 경로 C:\Users\wjddl\OneDrive\바탕 화면 -> 얜 절대경로, 구체적인 위치를 알 수 0
#위치 확인 방법: 파일>속성
#파이썬에서는 C:\Users\wjddl\OneDrive\바탕 화면 => C:\\Users\\wjddl\\OneDrive\\바탕 화면
# \두번씩 해야함.
# save 할 때도 수미상관이루어서 wb.save("C:\\Users\\wjddl\\OneDrive\\원하는 파일명") 이렇게 하면 됨.
except:                                                #기존에  파일이 없다면 (에러가 발생하면)
    wb = openpyxl.Workbook()                            #workbook을 새로 열고
    sheet =wb.active                                    #이경우엔 헤더 달아줘야함
    sheet.append(["제목", "언론사"])
    print("새로운 파일을 만들었습니다")
page = 1
for page in range (1,30,10):
    #데이터 요청을 반복 (10단위 페이지) / 요청할 데이터는 URL보고 파악하면 됨
    raw = requests.get("https://search.naver.com/search.naver?&where=news&query="+keyword+"&start="+str(page),
                       headers={"User-Agent":"Mozilla/5.0"})
    html = BeautifulSoup(raw.text, 'html.parser')

    #컨테이너: ul.type01 > li
    #기사제목: a._sp_each_title
    #언론사: span._sp_each_source

    #1. 컨테이너 수집
    articles = html.select("ul.type01 > li")

    #2. 기사 데이터 수집
    #title = articles[0].select_one("a._sp_each_title").text
    #source = articles[0].select_one("span._sp_each_source").text

    #print(title,source)
    #3. 반복하기
    for ar in articles:   #한 페이지 안에 데이터 수집 반복
          title = ar.select_one("a._sp_each_title").text
          source = ar.select_one("span._sp_each_source").text

          print(title, source)
          sheet.append([title, source])
wb.save("navernews.xlsx")